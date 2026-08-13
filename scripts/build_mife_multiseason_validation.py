#!/usr/bin/env python3
"""Build a de-identified, multi-season MIFE field-validation event inventory.

The raw workbooks are intentionally kept outside git because they contain GPS
coordinates and names. This script reads designated primary sheets only and
emits orchard/block-date event summaries. It does not modify model parameters.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class SourceSpec:
    filename: str
    sheet: str
    season: str
    region_group: str
    site_col: str
    block_col: str | None
    date_col: str
    tree_col: str
    sample_col: str
    organism_cols: tuple[str, ...]
    number_cols: tuple[str, ...]
    bbch_col: str | None


SOURCES = (
    SourceSpec("2023-2024 Dropsheet Sampling Final v2(1).xlsx", "Final", "2023-24", "mixed",
               "Farm", "Block", "CollectionDate", "Tree", "DropsheetSample",
               ("Consolidated_2026", "Insect Group"), ("Number",), None),
    SourceSpec("2024-2025 Drop Sheet Coates(2).xlsx", "Coates C1-1231", "2024-25", "nnsw",
               "Site", None, "phenology_date", "DT Code", "container",
               ("Syed ID",), ("Syed No.",), "BBCH:  majority of the tree"),
    SourceSpec("2024-2025 Drop Sheet Fuller(2).xlsx", "Sheet1", "2024-25", "seq",
               "Site", None, "phenology_date", "DT code", "Container",
               ("Syed ID",), ("Syed No.",), "BBCH:  majority of the tree"),
    SourceSpec("2024-2025 Drop Sheet Themsen(2).xlsx", "Sheet1", "2024-25", "cq",
               "Orchard", "Block", "phenology_date", "DT Code", "container",
               ("Syed_ID",), ("number",), "BBCH:  majority of the tree"),
    SourceSpec("2025-2026 Drop Sheet Coates 1(2).xlsx", "Sheet1", "2025-26", "nnsw",
               "farm_name", "block_name", "phenology_date", "ds_tree_code", "sort_container_code",
               ("Syed_ID", "sort_insect"), ("Syed_no", "no_bugs"), "BBCH_value_maj"),
    SourceSpec("2025-2026 Drop Sheet Fuller(2).xlsx", "Sheet1", "2025-26", "seq",
               "farm_name", None, "phenology_date", "ds_tree_code", "sort_container_code",
               ("Syed_ID", "sort_insect"), ("Syed_no", "no_bugs"), "BBCH_value_maj"),
)

SOURCE_IDS = {
    spec.filename: f"raw_{spec.season.replace('-', '_')}_{spec.region_group}_{index:02d}"
    for index, spec in enumerate(SOURCES, 1)
}

SITE_ALIASES = {
    "cheale": "cheal", "cheal": "cheal", "woolimi": "wollemi",
    "wollemi": "wollemi", "moore": "moore_park", "moore park": "moore_park",
    "north nursery farm": "nnf", "northern nursery farm": "nnf",
}
NNSW = {"dorey", "knockrow", "malua"}
SEQ = {"blanco", "cheal", "jackson", "jmac", "nnf", "bundaberg"}
CQ = {"alloway", "moore_park", "wollemi", "welcome_creek", "winfield"}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value)).casefold()


def norm_site(value: Any) -> str:
    raw = re.sub(r"[^a-z0-9]+", " ", clean(value).casefold()).strip()
    return SITE_ALIASES.get(raw, raw.replace(" ", "_"))


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            pass
    return None


def as_count(value: Any) -> float:
    try:
        result = float(value)
        return result if result >= 0 else 0.0
    except (TypeError, ValueError):
        return 0.0


def classify_target(label: Any) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", clean(label).casefold()).strip()
    if not text:
        return "other"
    has_fsb = bool(re.search(r"\bfsb\b|fruit spotting|fruitspotting", text))
    has_bsb = bool(re.search(r"\bbsb\b|banana spotting", text))
    if (has_fsb and has_bsb) or "possibly" in text:
        return "ambiguous_fsb_bsb"
    if has_fsb:
        return "confirmed_fsb"
    if has_bsb:
        return "confirmed_bsb"
    return "other"


def choose(row: dict[str, Any], candidates: Iterable[str]) -> Any:
    for candidate in candidates:
        value = row.get(norm_header(candidate))
        if clean(value):
            return value
    return None


def rows_from_sheet(path: Path, sheet: str) -> Iterable[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    iterator = ws.iter_rows(values_only=True)
    headers = [norm_header(v) for v in next(iterator)]
    for values in iterator:
        yield {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}


def event_inventory(source_dir: Path) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for spec in SOURCES:
        path = source_dir / spec.filename
        if not path.exists():
            raise FileNotFoundError(f"Missing source workbook: {path}")
        for row in rows_from_sheet(path, spec.sheet):
            site = norm_site(row.get(norm_header(spec.site_col)))
            visit = as_date(row.get(norm_header(spec.date_col)))
            if not site or not visit:
                continue
            block = clean(row.get(norm_header(spec.block_col))) if spec.block_col else ""
            key = (spec.season, site, block.casefold(), visit.isoformat())
            region = "nnsw" if site in NNSW else "seq" if site in SEQ else "cq" if site in CQ else spec.region_group
            event = grouped.setdefault(key, {
                "season": spec.season,
                "region": region,
                "site_id": site, "block_id": block, "visit_date": visit.isoformat(),
                "source_id": SOURCE_IDS[spec.filename],
                "trees": set(), "samples": set(), "bbch": [],
                "confirmed_fsb_count": 0.0, "confirmed_bsb_count": 0.0,
                "ambiguous_fsb_bsb_count": 0.0, "organism_rows": 0,
            })
            tree = clean(row.get(norm_header(spec.tree_col)))
            sample = clean(row.get(norm_header(spec.sample_col)))
            if tree:
                event["trees"].add(tree)
            if sample:
                event["samples"].add(sample)
            if spec.bbch_col:
                bbch_text = clean(row.get(norm_header(spec.bbch_col)))
                match = re.search(r"(\d{2})", bbch_text)
                if match:
                    event["bbch"].append(int(match.group(1)))
            label = choose(row, spec.organism_cols)
            if clean(label):
                event["organism_rows"] += 1
            target_class = classify_target(label)
            count = as_count(choose(row, spec.number_cols))
            target_field = {
                "confirmed_fsb": "confirmed_fsb_count",
                "confirmed_bsb": "confirmed_bsb_count",
                "ambiguous_fsb_bsb": "ambiguous_fsb_bsb_count",
            }.get(target_class)
            if target_field:
                event[target_field] += count or 1.0

    output = []
    for index, event in enumerate(sorted(grouped.values(), key=lambda x: (x["visit_date"], x["site_id"], x["block_id"])), 1):
        fsb = event.pop("confirmed_fsb_count")
        bsb = event.pop("confirmed_bsb_count")
        ambiguous = event.pop("ambiguous_fsb_bsb_count")
        trees = event.pop("trees")
        samples = event.pop("samples")
        bbch = event.pop("bbch")
        event.update({
            "event_id": f"MIFE-FV-{index:03d}",
            "recovered_tree_units": len(trees),
            "recovered_sample_units": len(samples),
            "median_bbch": median(bbch) if bbch else "",
            "confirmed_fsb_count": int(fsb) if fsb.is_integer() else fsb,
            "confirmed_bsb_count": int(bsb) if bsb.is_integer() else bsb,
            "ambiguous_fsb_bsb_count": int(ambiguous) if ambiguous.is_integer() else ambiguous,
            "confirmed_target_detected": int((fsb + bsb) > 0),
            "any_target_or_ambiguous_detected": int((fsb + bsb + ambiguous) > 0),
            "effort_note": "Recovered sample units; not necessarily planned sampling effort",
            "absence_note": "0 means no target row among recovered records, not confirmed orchard absence",
        })
        output.append(event)
    return output


def management_inventory(source_dir: Path) -> list[dict[str, Any]]:
    path = source_dir / "Management 2023-2026(2).xlsx"
    rows = list(rows_from_sheet(path, "All Mgmnt"))
    output = []
    for row in rows:
        site = norm_site(choose(row, ("Farm", "Site", "farm_name")))
        applied = as_date(choose(row, ("Date", "Application date", "date_applied")))
        if not site or not applied:
            continue
        reason = clean(choose(row, ("Reason", "Reason for application", "Target")))
        product = clean(choose(row, ("Chemistry", "Product", "Chemical", "Product name")))
        output.append({
            "site_id": site, "application_date": applied.isoformat(),
            "explicit_fsb_reason": int(bool(re.search(r"\bfsb\b|fruit[ -]?spot", reason, re.I))),
            "reason_recorded": int(bool(reason)), "product_recorded": int(bool(product)),
        })
    return sorted(output, key=lambda x: (x["application_date"], x["site_id"]))


def attach_management(events: list[dict[str, Any]], management: list[dict[str, Any]]) -> None:
    by_site: dict[str, list[tuple[date, int]]] = defaultdict(list)
    for row in management:
        by_site[row["site_id"]].append((date.fromisoformat(row["application_date"]), row["explicit_fsb_reason"]))
    for event in events:
        visit = date.fromisoformat(event["visit_date"])
        prior = [(visit - d).days for d, explicit in by_site[event["site_id"]] if explicit and 0 <= (visit - d).days <= 90]
        days = min(prior) if prior else None
        event["days_since_explicit_fsb_management"] = "" if days is None else days
        event["fsb_management_window"] = (
            "none_within_90d" if days is None else "0-7d" if days <= 7 else "8-21d" if days <= 21 else "22-90d"
        )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"No rows to write: {path}")
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def qa_rows(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in events:
        groups[(row["season"], row["region"])].append(row)
    output = []
    for (season, region), rows in sorted(groups.items()):
        output.append({
            "season": season, "region": region, "events": len(rows),
            "first_date": min(r["visit_date"] for r in rows),
            "last_date": max(r["visit_date"] for r in rows),
            "events_with_recovered_samples": sum(r["recovered_sample_units"] > 0 for r in rows),
            "confirmed_target_positive_events": sum(r["confirmed_target_detected"] for r in rows),
            "ambiguous_positive_events": sum(r["ambiguous_fsb_bsb_count"] > 0 for r in rows),
            "confirmed_fsb_total": sum(r["confirmed_fsb_count"] for r in rows),
            "confirmed_bsb_total": sum(r["confirmed_bsb_count"] for r in rows),
            "ambiguous_total": sum(r["ambiguous_fsb_bsb_count"] for r in rows),
        })
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=Path("data/validation"))
    args = parser.parse_args()
    events = event_inventory(args.source_dir)
    management = management_inventory(args.source_dir)
    attach_management(events, management)
    write_csv(args.output_dir / "MIFE_v0.1_multiseason_event_inventory.csv", events)
    write_csv(args.output_dir / "MIFE_v0.1_multiseason_management_deidentified.csv", management)
    write_csv(Path("data/output/MIFE_v0.1_multiseason_validation_QA.csv"), qa_rows(events))
    print(f"Wrote {len(events)} de-identified events and {len(management)} management records")


if __name__ == "__main__":
    main()
